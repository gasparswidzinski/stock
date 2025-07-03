from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QLabel, QPushButton, QComboBox, QLineEdit, 
    QHBoxLayout, QSpinBox, QTableWidget, QTableWidgetItem, QFileDialog, QMessageBox
)
from PySide6.QtCore import Qt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class BusquedaAvanzadaDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("🔍 Búsqueda Avanzada de Componentes")
        self.resize(900, 600)

        layout = QVBoxLayout(self)

        # Filtros
        f_layout = QHBoxLayout()
        layout.addLayout(f_layout)

        self.combo_etiqueta = QComboBox()
        self.combo_etiqueta.addItem("(Cualquiera)")
        etiquetas = self.db.obtener_todas_etiquetas()
        self.combo_etiqueta.addItems(etiquetas)
        f_layout.addWidget(QLabel("Etiqueta:"))
        f_layout.addWidget(self.combo_etiqueta)

        self.edit_proveedor = QLineEdit()
        f_layout.addWidget(QLabel("Proveedor contiene:"))
        f_layout.addWidget(self.edit_proveedor)

        self.spin_min = QSpinBox()
        self.spin_min.setMinimum(0)
        self.spin_min.setMaximum(999999)
        f_layout.addWidget(QLabel("Stock ≥"))
        f_layout.addWidget(self.spin_min)

        self.spin_max = QSpinBox()
        self.spin_max.setMinimum(0)
        self.spin_max.setMaximum(999999)
        self.spin_max.setValue(999999)
        f_layout.addWidget(QLabel("Stock ≤"))
        f_layout.addWidget(self.spin_max)

        # Tabla de resultados
        self.tabla = QTableWidget()
        self.tabla.setColumnCount(5)
        self.tabla.setHorizontalHeaderLabels([
            "ID", "Nombre", "Cantidad", "Proveedor", "Etiquetas"
        ])
        self.tabla.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tabla.setSelectionBehavior(QTableWidget.SelectRows)
        layout.addWidget(self.tabla)

        # Botones
        btn_buscar = QPushButton("Buscar")
        btn_buscar.clicked.connect(self._buscar)

        btn_exportar = QPushButton("Exportar resultados a Excel")
        btn_exportar.clicked.connect(self._exportar_excel)

        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.clicked.connect(self.accept)

        b_layout = QHBoxLayout()
        b_layout.addWidget(btn_buscar)
        b_layout.addWidget(btn_exportar)
        b_layout.addStretch()
        b_layout.addWidget(btn_cerrar)
        layout.addLayout(b_layout)

    def _buscar(self):
        etiqueta = self.combo_etiqueta.currentText()
        proveedor = self.edit_proveedor.text().strip().lower()
        min_stock = self.spin_min.value()
        max_stock = self.spin_max.value()

        c = self.db.conn.cursor()
        c.execute("""
            SELECT c.id, c.nombre, c.cantidad, c.proveedor,
                GROUP_CONCAT(e.nombre, ', ') as etiquetas
            FROM componentes c
            LEFT JOIN componentes_etiquetas ce ON c.id = ce.componente_id
            LEFT JOIN etiquetas e ON ce.etiqueta_id = e.id
            GROUP BY c.id
            ORDER BY c.nombre COLLATE NOCASE
        """)
        resultados = []
        for r in c.fetchall():
            # Filtros
            if etiqueta != "(Cualquiera)":
                etiquetas = r["etiquetas"].split(", ") if r["etiquetas"] else []
                if etiqueta not in etiquetas:
                    continue
            if proveedor and (not r["proveedor"] or proveedor not in r["proveedor"].lower()):
                continue
            if r["cantidad"] < min_stock or r["cantidad"] > max_stock:
                continue
            resultados.append(r)

        self._poblar_tabla(resultados)

    def _poblar_tabla(self, datos):
        self.tabla.setRowCount(0)
        for r in datos:
            row = self.tabla.rowCount()
            self.tabla.insertRow(row)
            self.tabla.setItem(row, 0, QTableWidgetItem(str(r["id"])))
            self.tabla.setItem(row, 1, QTableWidgetItem(r["nombre"]))
            self.tabla.setItem(row, 2, QTableWidgetItem(str(r["cantidad"])))
            self.tabla.setItem(row, 3, QTableWidgetItem(r["proveedor"] or "-"))
            self.tabla.setItem(row, 4, QTableWidgetItem(r["etiquetas"] or "-"))
        self.tabla.resizeColumnsToContents()

    def _exportar_excel(self):
        if self.tabla.rowCount() == 0:
            QMessageBox.warning(self, "Exportar", "No hay resultados para exportar.")
            return

        ruta, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar resultados como Excel",
            "resultados_busqueda.xlsx",
            "Archivos Excel (*.xlsx)"
        )
        if not ruta:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados búsqueda"

        headers = ["ID", "Nombre", "Cantidad", "Proveedor", "Etiquetas"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row in range(self.tabla.rowCount()):
            for col in range(self.tabla.columnCount()):
                ws.cell(row=row+2, column=col+1, value=self.tabla.item(row, col).text())

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 2

        wb.save(ruta)
        QMessageBox.information(self, "Exportar", f"Archivo guardado:\n{ruta}")
