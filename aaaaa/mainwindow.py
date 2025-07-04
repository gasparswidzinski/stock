from PySide6.QtWidgets import (
    QMainWindow, QLineEdit, QDialog, QVBoxLayout, QTableWidget,
    QTableWidgetItem, QPushButton, QHBoxLayout, QListWidget, QMessageBox, QInputDialog,QLabel, QDialogButtonBox
)
from PySide6.QtGui import QAction
from PySide6.QtGui import QKeySequence, QShortcut
from PySide6.QtCore import Qt, QSettings, Slot
from db import DBManager
from dialogs import ItemDialog
from docks import TagDockWidget
from proyectos import ProyectoWidget
from proyectos_dialog import ProyectosDialog
from dashboard import DashboardWidget
from PySide6.QtGui import QIcon
from PySide6 import QtGui
from historial_dialog import HistorialDialog
from PySide6.QtWidgets import QMessageBox
import os
import shutil
from datetime import datetime
from estadisticas_dialog import EstadisticasDialog
from PySide6.QtWidgets import QFileDialog
import csv
from PySide6.QtGui import QPixmap
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from busqueda_avanzada_dialog import BusquedaAvanzadaDialog
import smtplib
from email.mime.text import MIMEText
from PySide6.QtWidgets import QGraphicsOpacityEffect
from PySide6.QtCore import QPropertyAnimation, QEasingCurve
from configuracion_dialog import ConfiguracionDialog
from utils import cargar_tema



class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Control de Stock de Componentes")
        self.db = DBManager()
        self.tema_actual = "claro"
        self.docks = {}
        self._build_actions()
        self._restore_ui_state()
        self._abrir_etiquetas_existentes()
        self._alertar_stock_bajo()
        self._verificar_integridad_datos()
        self._mostrar_dashboard()
        self._aplicar_configuracion_visual()
        effect = QGraphicsOpacityEffect(self)
        self.setGraphicsEffect(effect)

        anim = QPropertyAnimation(effect, b"opacity")
        anim.setDuration(500)
        anim.setStartValue(0)
        anim.setEndValue(1)
        anim.setEasingCurve(QEasingCurve.OutCubic)
        anim.start()
        self._fade_animation = anim

    def _build_actions(self):

        menubar = self.menuBar()
        menu_archivo = menubar.addMenu("Archivo")

        act_agregar = QAction(QIcon("icons/add.png"), "Agregar componente", self)
        act_agregar.triggered.connect(self._on_agregar_componente)
        menu_archivo.addAction(act_agregar)

        act_salir = QAction("Salir", self)
        act_salir.triggered.connect(self.close)
        menu_archivo.addAction(act_salir)

        toolbar = self.addToolBar("Principal")
        toolbar.setObjectName("MainToolbar")

        self.edit_buscar = QLineEdit()
        self.edit_buscar.setPlaceholderText("Buscar...")
        self.edit_buscar.returnPressed.connect(self._on_buscar_global)
        toolbar.addWidget(self.edit_buscar)

        btn_buscar = QAction(QIcon("icons/inventory.png"), "Buscar", self)
        btn_buscar.triggered.connect(self._on_buscar_global)
        toolbar.addAction(btn_buscar)

        shortcut_buscar = QShortcut(QKeySequence("Ctrl+F"), self)
        shortcut_buscar.activated.connect(self._mostrar_dialogo_busqueda)

        menu_ver = menubar.addMenu("Menú")

        menu_tema = menubar.addMenu("Tema")

        act_tema_claro = QAction("Tema Claro", self)
        act_tema_claro.triggered.connect(lambda: self._cargar_tema("tema_claro.qss"))
        menu_tema.addAction(act_tema_claro)

        act_tema_oscuro = QAction("Tema Oscuro", self)
        act_tema_oscuro.triggered.connect(lambda: self._cargar_tema("tema_oscuro.qss"))
        menu_tema.addAction(act_tema_oscuro)

        act_ver_etiquetas = QAction(QIcon("icons/folder.png"), "Administrar etiquetas", self)
        act_ver_etiquetas.triggered.connect(self._on_administrar_etiquetas)
        menu_ver.addAction(act_ver_etiquetas)

        act_activar_oscuro = QAction("Activar tema oscuro", self)
        act_activar_oscuro.triggered.connect(self.aplicar_tema_oscuro)
        menu_ver.addAction(act_activar_oscuro)

        act_volver_claro = QAction("Volver al tema claro", self)
        act_volver_claro.triggered.connect(self.aplicar_tema_claro)
        menu_ver.addAction(act_volver_claro)

        act_nuevo_proyecto = QAction(QIcon("icons/folder.png"), "Nuevo Proyecto", self)
        act_nuevo_proyecto.triggered.connect(self._nuevo_proyecto)
        menu_ver.addAction(act_nuevo_proyecto)

        act_ver_proyectos = QAction(QIcon("icons/folder.png"), "Ver proyectos guardados", self)
        act_ver_proyectos.triggered.connect(self._abrir_dialogo_proyectos)
        menu_ver.addAction(act_ver_proyectos)

        act_importar = QAction(QIcon("icons/inventory.png"), "Importar proyecto desde CSV", self)
        act_importar.triggered.connect(self._importar_proyecto_csv)
        menu_ver.addAction(act_importar)

        # Atajos de teclado
        shortcut_nuevo = QShortcut(QKeySequence("Ctrl+N"), self)
        shortcut_nuevo.activated.connect(self._on_agregar_componente)

        shortcut_proyecto = QShortcut(QKeySequence("Ctrl+P"), self)
        shortcut_proyecto.activated.connect(self._nuevo_proyecto)

        shortcut_stock = QShortcut(QKeySequence("F3"), self)
        shortcut_stock.activated.connect(self._entrada_rapida_stock)

        shortcut_todo = QShortcut(QKeySequence("F4"), self)
        shortcut_todo.activated.connect(self._mostrar_todo_stock)

        shortcut_historial = QShortcut(QKeySequence("Ctrl+H"), self)
        shortcut_historial.activated.connect(self._abrir_dialogo_proyectos)
        
        act_ver_historial = QAction("Ver historial de movimientos", self)
        act_ver_historial.triggered.connect(self._abrir_historial)
        menu_ver.addAction(act_ver_historial)
        shortcut_historial = QShortcut(QKeySequence("Ctrl+M"), self)
        shortcut_historial.activated.connect(self._abrir_historial)
        
        act_estadisticas = QAction("Ver estadísticas avanzadas", self)
        act_estadisticas.triggered.connect(self._abrir_estadisticas)
        menu_ver.addAction(act_estadisticas)
        
        act_exportar_inventario = QAction("Exportar inventario a CSV", self)
        act_exportar_inventario.triggered.connect(self._exportar_inventario_csv)
        menu_ver.addAction(act_exportar_inventario)

        act_exportar_historial = QAction("Exportar historial a CSV", self)
        act_exportar_historial.triggered.connect(self._exportar_historial_csv)
        menu_ver.addAction(act_exportar_historial)
        
        act_exportar_excel = QAction("Exportar inventario a Excel", self)
        act_exportar_excel.triggered.connect(self._exportar_inventario_excel)
        menu_ver.addAction(act_exportar_excel)
        
        act_exportar_historial_excel = QAction("Exportar historial a Excel", self)
        act_exportar_historial_excel.triggered.connect(self._exportar_historial_excel)
        menu_ver.addAction(act_exportar_historial_excel)
        
        act_busqueda_avanzada = QAction("Búsqueda avanzada", self)
        act_busqueda_avanzada.triggered.connect(self._abrir_busqueda_avanzada)
        menu_ver.addAction(act_busqueda_avanzada)
        
        act_configuracion = QAction("Configuración visual", self)
        act_configuracion.triggered.connect(self._abrir_configuracion)
        menu_ver.addAction(act_configuracion)



        
    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Insert:
            self._on_agregar_componente()
        else:
            super().keyPressEvent(event)   

    
    def _mostrar_dialogo_busqueda(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Buscar componente")
        layout = QVBoxLayout(dlg)

        label = QLabel("Ingrese el texto a buscar:")
        layout.addWidget(label)

        input_buscar = QLineEdit()
        layout.addWidget(input_buscar)

        botones = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(botones)

        botones.accepted.connect(dlg.accept)
        botones.rejected.connect(dlg.reject)

        if dlg.exec():
            texto = input_buscar.text().strip()
            self._on_buscar_global(texto)

    @Slot()
    def _on_agregar_componente(self):
        dlg = ItemDialog(self.db, parent=self)
        dlg.componente_guardado.connect(self.on_componente_guardado)
        dlg.exec()

    @Slot(int)
    def on_componente_guardado(self, comp_id: int):
        etiquetas = self.db.obtener_etiquetas_por_componente(comp_id)
        for tag in etiquetas:
            if tag not in self.docks:
                dock = TagDockWidget(tag, self.db, parent=self)
                self.addDockWidget(Qt.RightDockWidgetArea, dock)
                self.docks[tag] = dock
            else:
                dock = self.docks[tag]
            dock.actualizar_lista()

        for tag, dock in list(self.docks.items()):
            if tag not in etiquetas:
                dock.actualizar_lista()
                if not self.db.obtener_componentes_por_etiqueta(tag):
                    dock.close()
                    del self.docks[tag]

    @Slot(int)
    def on_componente_eliminado(self, comp_id: int):
        for tag, dock in list(self.docks.items()):
            items = self.db.obtener_componentes_por_etiqueta(tag)
            if not items:
                dock.close()
                del self.docks[tag]
            else:
                dock.actualizar_lista()

    @Slot()
    @Slot(str)
    def _on_buscar_global(self, texto=None):
        if texto is None:
            texto = self.edit_buscar.text().strip().lower()
        else:
            texto = texto.lower()
        if not texto:
            return

        pattern = f"%{texto}%"
        c = self.db.conn.cursor()
        c.execute('''
        SELECT DISTINCT c.* FROM componentes c
        LEFT JOIN componentes_etiquetas ce ON c.id=ce.componente_id
        LEFT JOIN etiquetas e ON ce.etiqueta_id=e.id
        WHERE LOWER(c.nombre) LIKE ?
        OR LOWER(c.valor) LIKE ?
        OR LOWER(c.descripcion) LIKE ?
        OR LOWER(c.ubicacion) LIKE ?
        OR LOWER(c.proveedor) LIKE ?
        OR LOWER(e.nombre) LIKE ?
        ORDER BY c.nombre COLLATE NOCASE;
        ''', (pattern,) * 6)
        resultados = [dict(r) for r in c.fetchall()]

        dlg = QDialog(self)
        dlg.setWindowTitle(f"Resultados de búsqueda: '{texto}'")
        v = QVBoxLayout(dlg)
        tabla = QTableWidget()
        tabla.setColumnCount(3)
        tabla.setHorizontalHeaderLabels(["ID", "Nombre", "Etiquetas"])
        tabla.setEditTriggers(QTableWidget.NoEditTriggers)
        for row_data in resultados:
            row = tabla.rowCount()
            tabla.insertRow(row)
            tabla.setItem(row, 0, QTableWidgetItem(str(row_data["id"])))
            tabla.setItem(row, 1, QTableWidgetItem(row_data["nombre"]))
            etiquetas = self.db.obtener_etiquetas_por_componente(row_data["id"])
            tabla.setItem(row, 2, QTableWidgetItem(", ".join(etiquetas)))
        tabla.resizeColumnsToContents()
        v.addWidget(tabla)

        btn_editar = QPushButton("Editar seleccionado")
        def editar_sel():
            sel = tabla.selectedItems()
            if not sel:
                return
            comp_id = int(tabla.item(sel[0].row(), 0).text())
            dlg2 = ItemDialog(self.db, parent=self, comp_id=comp_id)
            dlg2.componente_guardado.connect(self.on_componente_guardado)
            dlg2.exec()
            dlg.accept()
            self._on_buscar_global()
        btn_editar.clicked.connect(editar_sel)

        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.clicked.connect(dlg.accept)
        h = QHBoxLayout()
        h.addWidget(btn_editar)
        h.addWidget(btn_cerrar)
        v.addLayout(h)
        dlg.resize(500, 300)
        dlg.exec()

    @Slot()
    def _on_administrar_etiquetas(self):
        etiquetas = self.db.obtener_todas_etiquetas()
        dlg = QDialog(self)
        dlg.setWindowTitle("Administrar etiquetas")
        v = QVBoxLayout(dlg)
        listw = QListWidget()
        listw.addItems(etiquetas)
        v.addWidget(listw)

        def renombrar():
            sel = listw.selectedItems()
            if not sel: return
            old = sel[0].text()
            nuevo, ok = QInputDialog.getText(self, "Renombrar etiqueta", f"Nuevo nombre para '{old}':")
            if ok and nuevo.strip():
                c = self.db.conn.cursor()
                try:
                    c.execute("UPDATE etiquetas SET nombre=? WHERE nombre=?", (nuevo.strip(), old))
                    self.db.conn.commit()
                    if old in self.docks:
                        dock = self.docks.pop(old)
                        dock.etiqueta = nuevo.strip()
                        dock.setWindowTitle(nuevo.strip())
                        dock.setObjectName(f"Dock_{nuevo.strip()}")
                        self.docks[nuevo.strip()] = dock
                    sel[0].setText(nuevo.strip())
                except:
                    QMessageBox.warning(self, "Error", "Ya existe una etiqueta con ese nombre.")

        def eliminar():
            sel = listw.selectedItems()
            if not sel: return
            tag = sel[0].text()
            resp = QMessageBox.question(self, "Confirmar", f"Eliminar etiqueta '{tag}' de todos los componentes?")
            if resp == QMessageBox.Yes:
                c = self.db.conn.cursor()
                c.execute("DELETE FROM etiquetas WHERE nombre=?", (tag,))
                self.db.conn.commit()
                if tag in self.docks:
                    dock = self.docks.pop(tag)
                    dock.close()
                listw.takeItem(listw.row(sel[0]))

        btn_renombrar = QPushButton("Renombrar")
        btn_renombrar.clicked.connect(renombrar)
        btn_eliminar = QPushButton("Eliminar")
        btn_eliminar.clicked.connect(eliminar)
        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.clicked.connect(dlg.accept)
        h = QHBoxLayout()
        h.addWidget(btn_renombrar)
        h.addWidget(btn_eliminar)
        h.addWidget(btn_cerrar)
        v.addLayout(h)
        dlg.resize(300, 400)
        dlg.exec()

    def closeEvent(self, event):

        # Guardar geometría y estado
        settings = QSettings("MiInventario", "AppStock")
        settings.setValue("geometry", self.saveGeometry())
        settings.setValue("windowState", self.saveState())

        # Backup automático
        try:
            backups_dir = os.path.join(os.path.dirname(__file__), "backups")
            if not os.path.exists(backups_dir):
                os.makedirs(backups_dir)

            fecha = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            backup_file = os.path.join(backups_dir, f"stock_backup_{fecha}.db")

            shutil.copyfile(self.db.db_path, backup_file)
        except Exception as e:
            QMessageBox.warning(
                self,
                "Backup fallido",
                f"No se pudo crear el backup automático:\n{e}"
            )

        super().closeEvent(event)

    def _restore_ui_state(self):
        settings = QSettings("MiInventario", "AppStock")
        geom = settings.value("geometry")
        if geom:
            self.restoreGeometry(geom)
        state = settings.value("windowState")
        if state:
            self.restoreState(state)
    
    def aplicar_tema_oscuro(self):
       self.setStyleSheet("""
        * {
            font-family: 'Segoe UI', sans-serif;
            font-size: 10.5pt;
        }
        QMainWindow {
            background-color: #202124;
            color: #e8eaed;
        }
        QMenuBar, QMenu {
            background-color: #303134;
            color: #e8eaed;
        }
        QMenu::item:selected {
            background-color: #3c4043;
        }
        QToolBar {
            background-color: #2a2a2d;
        }
        QDialog, QDockWidget, QWidget {
            background-color: #202124;
            color: #e8eaed;
        }
        QLabel, QListWidget, QLineEdit, QTextEdit, QSpinBox, QTableWidget, QPushButton {
            background-color: #292a2d;
            color: #e8eaed;
            border: 1px solid #3c4043;
            padding: 2px;
        }
        QLineEdit:focus, QTextEdit:focus, QSpinBox:focus {
            border: 1px solid #8ab4f8;
        }
        QPushButton {
            background-color: #3c4043;
            border-radius: 4px;
        }
        QPushButton:hover {
            background-color: #5f6368;
        }
        QHeaderView::section {
            background-color: #3c4043;
            color: #e8eaed;
            padding: 4px;
        }
        QTableWidget {
            gridline-color: #444;
            selection-background-color: #5f6368;
            selection-color: white;
        }
    """)
    def aplicar_tema_claro(self):
        self.setStyleSheet("")
    
    def _abrir_etiquetas_existentes(self):
        etiquetas = self.db.obtener_todas_etiquetas()
        for tag in etiquetas:
            componentes = self.db.obtener_componentes_por_etiqueta(tag)
            if componentes:
                if tag not in self.docks:
                    dock = TagDockWidget(tag, self.db, parent=self)
                    self.addDockWidget(Qt.RightDockWidgetArea, dock)
                    self.docks[tag] = dock
    
    
    def _nuevo_proyecto(self):
        widget = ProyectoWidget(self.db, self)
        self.setCentralWidget(widget)
    
    
    def _dialogo_nombre_proyecto(self, titulo):
        from PySide6.QtWidgets import QInputDialog
        return QInputDialog.getText(self, titulo, "Nombre del proyecto:")
    
    def _alertar_stock_bajo(self):
        c = self.db.conn.cursor()
        c.execute("""
            SELECT nombre, cantidad, stock_minimo
            FROM componentes
            WHERE stock_minimo IS NOT NULL AND cantidad < stock_minimo
            ORDER BY nombre COLLATE NOCASE
        """)
        resultados = c.fetchall()
        if resultados:
            lista = "\n".join(
                f"{r['nombre']} (Stock: {r['cantidad']}, Mínimo: {r['stock_minimo']})"
                for r in resultados
            )
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.warning(
                self,
                "⚠️ Alerta de stock bajo",
                f"Los siguientes componentes tienen stock por debajo del mínimo:\n\n{lista}"
            )
            #self._enviar_alerta_email(resultados)

    def _entrada_rapida_stock(self):
        from PySide6.QtWidgets import QInputDialog, QMessageBox
        comp_id_str, ok = QInputDialog.getText(self, "Entrada rápida de stock", "ID del componente:")
        if not ok or not comp_id_str.strip():
            return
        try:
            comp_id = int(comp_id_str.strip())
        except ValueError:
            QMessageBox.warning(self, "Error", "ID inválido.")
            return
        c = self.db.conn.cursor()
        c.execute("SELECT nombre, cantidad FROM componentes WHERE id=?", (comp_id,))
        row = c.fetchone()
        if not row:
            QMessageBox.warning(self, "Error", "No se encontró el componente.")
            return
        cantidad_str, ok = QInputDialog.getText(
            self, "Cantidad a sumar",
            f"Componente: {row['nombre']} (Stock actual: {row['cantidad']})\nCantidad a sumar:"
        )
        if not ok or not cantidad_str.strip():
            return
        try:
            cantidad = int(cantidad_str.strip())
            if cantidad <= 0:
                raise ValueError()
        except ValueError:
            QMessageBox.warning(self, "Error", "Cantidad inválida.")
            return
        c.execute("UPDATE componentes SET cantidad = cantidad + ? WHERE id=?", (cantidad, comp_id))
        self.db.conn.commit()
        c.execute("""
            INSERT INTO historial (componente_id, accion, cantidad, fecha_hora, descripcion)
            VALUES (?, 'Entrada Stock', ?, datetime('now'), 'Entrada rápida desde F3')
        """, (comp_id, cantidad))
        self.db.conn.commit()
        QMessageBox.information(
            self, "Stock actualizado",
            f"Nuevo stock de '{row['nombre']}': {row['cantidad'] + cantidad}"
        )
    
    def _mostrar_todo_stock(self):
        from PySide6.QtWidgets import (
            QDialog, QVBoxLayout, QTableWidget, QTableWidgetItem,
            QPushButton, QHBoxLayout, QMessageBox
        )
        from PySide6.QtCore import Qt
        from dialogs import ItemDialog

        dlg = QDialog(self)
        dlg.setWindowTitle("Inventario completo")
        layout = QVBoxLayout(dlg)

        tabla = QTableWidget()
        tabla.setSelectionBehavior(QTableWidget.SelectRows)
        tabla.setSelectionMode(QTableWidget.ExtendedSelection)
        tabla.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(tabla)

        modo_vista = self._obtener_vista_stock()

        def cargar_datos():
            tabla.setRowCount(0)
            c = self.db.conn.cursor()
            c.execute("""
                SELECT id, nombre, cantidad, stock_minimo, ubicacion, proveedor
                FROM componentes
            """)
            componentes = []
            for r in c.fetchall():
                etiquetas = self.db.obtener_etiquetas_por_componente(r["id"])
                etiqueta = etiquetas[0] if etiquetas else "(Sin etiqueta)"
                componentes.append({
                    "id": r["id"],
                    "etiqueta": etiqueta,
                    "nombre": r["nombre"],
                    "cantidad": r["cantidad"],
                    "stock_minimo": r["stock_minimo"],
                    "ubicacion": r["ubicacion"],
                    "proveedor": r["proveedor"],
                })
            componentes.sort(key=lambda x: (x["etiqueta"].lower(), x["nombre"].lower()))

            # Configurar columnas según modo
            if modo_vista == "compacta":
                columnas = ["ID", "Nombre", "Cantidad"]
            else:
                columnas = ["ID", "Etiqueta", "Nombre", "Cantidad", "Stock mínimo", "Ubicación", "Proveedor"]

            tabla.setColumnCount(len(columnas))
            tabla.setHorizontalHeaderLabels(columnas)

            for r in componentes:
                row = tabla.rowCount()
                tabla.insertRow(row)
                if modo_vista == "compacta":
                    tabla.setItem(row, 0, QTableWidgetItem(str(r["id"])))
                    tabla.setItem(row, 1, QTableWidgetItem(r["nombre"]))
                    tabla.setItem(row, 2, QTableWidgetItem(str(r["cantidad"])))
                else:
                    tabla.setItem(row, 0, QTableWidgetItem(str(r["id"])))
                    tabla.setItem(row, 1, QTableWidgetItem(r["etiqueta"]))
                    tabla.setItem(row, 2, QTableWidgetItem(r["nombre"]))
                    tabla.setItem(row, 3, QTableWidgetItem(str(r["cantidad"])))
                    tabla.setItem(row, 4, QTableWidgetItem(str(r["stock_minimo"]) if r["stock_minimo"] else "-"))
                    tabla.setItem(row, 5, QTableWidgetItem(r["ubicacion"] or "-"))
                    tabla.setItem(row, 6, QTableWidgetItem(r["proveedor"] or "-"))

            tabla.resizeColumnsToContents()

            # 🔹 Colorear filas con stock bajo (solo modo detalle)
            if modo_vista == "detalle":
                for row in range(tabla.rowCount()):
                    cantidad = int(tabla.item(row, 3).text())
                    stock_min_text = tabla.item(row, 4).text()
                    stock_min = int(stock_min_text) if stock_min_text != "-" else None

                    if stock_min is not None and cantidad < stock_min:
                        for col in range(tabla.columnCount()):
                            tabla.item(row, col).setBackground(Qt.red)
                            tabla.item(row, col).setForeground(Qt.white)

        cargar_datos()

        
        btn_cambiar_vista = QPushButton("Cambiar Vista")
        btn_editar = QPushButton("Editar seleccionado")
        btn_eliminar = QPushButton("Eliminar seleccionados")
        btn_cerrar = QPushButton("Cerrar")

        def cambiar_vista():
            nonlocal modo_vista
            if modo_vista == "detalle":
                modo_vista = "compacta"
            else:
                modo_vista = "detalle"
            self._guardar_vista_stock(modo_vista)
            cargar_datos()

        def editar():
            sel = tabla.selectedItems()
            if not sel:
                return
            comp_id = int(tabla.item(sel[0].row(), 0).text())
            dlg_edit = ItemDialog(self.db, self, comp_id=comp_id)
            dlg_edit.componente_guardado.connect(lambda _: cargar_datos())
            dlg_edit.exec()

        def eliminar():
            selected = tabla.selectionModel().selectedRows()
            if not selected:
                return
            resp = QMessageBox.question(
                dlg, "Confirmar eliminación",
                f"¿Seguro que deseas eliminar {len(selected)} componente(s)?"
            )
            if resp != QMessageBox.Yes:
                return
            for idx in sorted(selected, key=lambda x: x.row(), reverse=True):
                comp_id = int(tabla.item(idx.row(), 0).text())
                self.db.eliminar_componente(comp_id)
            cargar_datos()

        btn_cambiar_vista.clicked.connect(cambiar_vista)
        btn_editar.clicked.connect(editar)
        btn_eliminar.clicked.connect(eliminar)
        btn_cerrar.clicked.connect(dlg.accept)

        h = QHBoxLayout()
        h.addWidget(btn_cambiar_vista)
        h.addWidget(btn_editar)
        h.addWidget(btn_eliminar)
        h.addStretch()
        h.addWidget(btn_cerrar)
        layout.addLayout(h)

        dlg.resize(800, 500)
        dlg.exec()

        
    def _abrir_dialogo_proyectos(self):
        dlg = ProyectosDialog(self.db, self)
        dlg.exec()

    def _importar_proyecto_csv(self):
        from PySide6.QtWidgets import QFileDialog, QMessageBox
        import csv

        ruta, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar CSV",
            "",
            "Archivos CSV (*.csv)"
        )
        if not ruta:
            return

        componentes = []
        try:
            with open(ruta, newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                headers = next(reader)
                for row in reader:
                    if len(row) < 2:
                        continue
                    comp_id = int(row[0].strip())
                    cantidad = int(row[1].strip())
                    if cantidad <= 0:
                        continue
                    componentes.append({"id": comp_id, "cantidad": cantidad})
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo leer el archivo:\n{e}")
            return

        if not componentes:
            QMessageBox.warning(self, "Error", "El archivo no contiene datos válidos.")
            return

        
        from proyectos import ProyectoWidget
        widget = ProyectoWidget(self.db, self)
        widget.componentes = componentes
        widget._actualizar_tabla()
        self.setCentralWidget(widget)
    
    def _mostrar_dashboard(self):
        dashboard = DashboardWidget(self.db, self)
        self.setCentralWidget(dashboard)
    
    def _cargar_tema(self, nombre_archivo):
        import os
        ruta = os.path.join(os.path.dirname(__file__), nombre_archivo)
        with open(ruta, "r", encoding="utf-8") as f:
            self.setStyleSheet(f.read())

        if "oscuro" in nombre_archivo:
            self.tema_actual = "oscuro"
        else:
            self.tema_actual = "claro"
            
    def _abrir_historial(self):
        dlg = HistorialDialog(self.db, self)
        dlg.exec()

    def _verificar_integridad_datos(self):
        errores = []

        c = self.db.conn.cursor()

        # Verificar stock negativo
        c.execute("""
            SELECT id, nombre, cantidad
            FROM componentes
            WHERE cantidad < 0
            ORDER BY nombre COLLATE NOCASE
        """)
        negativos = c.fetchall()
        if negativos:
            for r in negativos:
                errores.append(
                    f"Stock negativo: {r['nombre']} (ID {r['id']}), cantidad: {r['cantidad']}"
                )

        # Si hay errores, mostrar alerta
        if errores:
            from PySide6.QtWidgets import QMessageBox
            mensaje = "\n".join(errores)
            QMessageBox.warning(
                self,
                "Problemas de integridad de datos",
                f"Se detectaron los siguientes problemas:\n\n{mensaje}"
            )

    def _abrir_estadisticas(self):
        dlg = EstadisticasDialog(self.db, self)
        dlg.exec()
    def _exportar_inventario_csv(self):
        

        ruta, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar inventario como CSV",
            "inventario.csv",
            "Archivos CSV (*.csv)"
        )
        if not ruta:
            return

        c = self.db.conn.cursor()
        c.execute("""
            SELECT c.id, c.nombre, c.cantidad, c.stock_minimo, c.ubicacion, c.proveedor,
                GROUP_CONCAT(e.nombre, ', ') as etiquetas
            FROM componentes c
            LEFT JOIN componentes_etiquetas ce ON c.id = ce.componente_id
            LEFT JOIN etiquetas e ON ce.etiqueta_id = e.id
            GROUP BY c.id
            ORDER BY c.nombre COLLATE NOCASE
        """)
        filas = c.fetchall()

        try:
            with open(ruta, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "ID", "Nombre", "Cantidad", "Stock mínimo",
                    "Ubicación", "Proveedor", "Etiquetas"
                ])
                for r in filas:
                    writer.writerow([
                        r["id"], r["nombre"], r["cantidad"], r["stock_minimo"],
                        r["ubicacion"], r["proveedor"], r["etiquetas"]
                    ])
            QMessageBox.information(
                self, "Exportación exitosa",
                f"Inventario exportado correctamente a:\n{ruta}"
            )
        except Exception as e:
            QMessageBox.warning(
                self, "Error al exportar",
                f"No se pudo exportar el inventario:\n{e}"
            )
    def _exportar_historial_csv(self):
        
        ruta, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar historial como CSV",
            "historial.csv",
            "Archivos CSV (*.csv)"
        )
        if not ruta:
            return

        c = self.db.conn.cursor()
        c.execute("""
            SELECT h.fecha_hora, h.accion, c.nombre, h.cantidad, h.descripcion
            FROM historial h
            LEFT JOIN componentes c ON h.componente_id = c.id
            ORDER BY h.fecha_hora DESC
        """)
        filas = c.fetchall()

        try:
            with open(ruta, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "Fecha y Hora", "Acción", "Componente",
                    "Cantidad", "Descripción"
                ])
                for r in filas:
                    writer.writerow([
                        r["fecha_hora"], r["accion"], r["nombre"],
                        r["cantidad"], r["descripcion"]
                    ])
            QMessageBox.information(
                self, "Exportación exitosa",
                f"Historial exportado correctamente a:\n{ruta}"
            )
        except Exception as e:
            QMessageBox.warning(
                self, "Error al exportar",
                f"No se pudo exportar el historial:\n{e}"
            )

    def _mostrar_detalle_componente(self, comp_id):
        c = self.db.conn.cursor()
        c.execute("SELECT * FROM componentes WHERE id=?", (comp_id,))
        r = c.fetchone()
        if not r:
            QMessageBox.warning(self, "Error", "No se encontró el componente.")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle(f"Detalle de '{r['nombre']}'")
        dlg.resize(500, 600)
        layout = QVBoxLayout(dlg)

        # Imagen
        imagen_path = r["imagen_path"]
        if imagen_path and os.path.isfile(imagen_path):
            pixmap = QPixmap(imagen_path).scaled(400, 400, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            lbl_img = QLabel()
            lbl_img.setPixmap(pixmap)
            lbl_img.setAlignment(Qt.AlignCenter)
            layout.addWidget(lbl_img)

        # Detalles
        detalles = f"""
    <b>ID:</b> {r['id']}<br>
    <b>Nombre:</b> {r['nombre']}<br>
    <b>Valor:</b> {r['valor'] or '-'}<br>
    <b>Cantidad:</b> {r['cantidad']}<br>
    <b>Stock mínimo:</b> {r['stock_minimo'] or '-'}<br>
    <b>Ubicación:</b> {r['ubicacion'] or '-'}<br>
    <b>Proveedor:</b> {r['proveedor'] or '-'}<br>
    <b>Descripción:</b><br>{r['descripcion'] or '-'}
    """
        lbl_detalle = QLabel(detalles)
        lbl_detalle.setTextFormat(Qt.RichText)
        lbl_detalle.setWordWrap(True)
        layout.addWidget(lbl_detalle)

        # Botón cerrar
        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.clicked.connect(dlg.accept)
        layout.addWidget(btn_cerrar)

        dlg.exec()
    
    def _exportar_inventario_excel(self):

        ruta, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar inventario como Excel",
            "inventario.xlsx",
            "Archivos Excel (*.xlsx)"
        )
        if not ruta:
            return

        c = self.db.conn.cursor()
        c.execute("""
            SELECT c.id, c.nombre, c.cantidad, c.stock_minimo, c.ubicacion, c.proveedor,
                GROUP_CONCAT(e.nombre, ', ') as etiquetas
            FROM componentes c
            LEFT JOIN componentes_etiquetas ce ON c.id = ce.componente_id
            LEFT JOIN etiquetas e ON ce.etiqueta_id = e.id
            GROUP BY c.id
            ORDER BY c.nombre COLLATE NOCASE
        """)
        filas = c.fetchall()

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Inventario"

            # Encabezados
            headers = [
                "ID", "Nombre", "Cantidad", "Stock mínimo",
                "Ubicación", "Proveedor", "Etiquetas"
            ]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Filas de datos
            for row_num, r in enumerate(filas, 2):
                ws.cell(row=row_num, column=1, value=r["id"])
                ws.cell(row=row_num, column=2, value=r["nombre"])
                ws.cell(row=row_num, column=3, value=r["cantidad"])
                ws.cell(row=row_num, column=4, value=r["stock_minimo"])
                ws.cell(row=row_num, column=5, value=r["ubicacion"])
                ws.cell(row=row_num, column=6, value=r["proveedor"])
                ws.cell(row=row_num, column=7, value=r["etiquetas"])

            # Ajuste de ancho
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(ruta)
            QMessageBox.information(
                self, "Exportación exitosa",
                f"Inventario exportado correctamente a:\n{ruta}"
            )

        except Exception as e:
            QMessageBox.warning(
                self, "Error al exportar",
                f"No se pudo exportar el inventario:\n{e}"
            )
    def _exportar_historial_excel(self):
       
        ruta, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar historial como Excel",
            "historial.xlsx",
            "Archivos Excel (*.xlsx)"
        )
        if not ruta:
            return

        c = self.db.conn.cursor()
        c.execute("""
            SELECT h.fecha_hora, h.accion, c.nombre, h.cantidad, h.descripcion
            FROM historial h
            LEFT JOIN componentes c ON h.componente_id = c.id
            ORDER BY h.fecha_hora DESC
        """)
        filas = c.fetchall()

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Historial"

            # Encabezados
            headers = ["Fecha y Hora", "Acción", "Componente", "Cantidad", "Descripción"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Filas de datos
            for row_num, r in enumerate(filas, 2):
                ws.cell(row=row_num, column=1, value=r["fecha_hora"])
                ws.cell(row=row_num, column=2, value=r["accion"])
                ws.cell(row=row_num, column=3, value=r["nombre"])
                ws.cell(row=row_num, column=4, value=r["cantidad"])
                ws.cell(row=row_num, column=5, value=r["descripcion"])

            # Ajuste de ancho
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = max_length + 2

            wb.save(ruta)
            QMessageBox.information(
                self, "Exportación exitosa",
                f"Historial exportado correctamente a:\n{ruta}"
            )

        except Exception as e:
            QMessageBox.warning(
                self, "Error al exportar",
                f"No se pudo exportar el historial:\n{e}"
            )
    def _abrir_busqueda_avanzada(self):
        
        dlg = BusquedaAvanzadaDialog(self.db, self)
        dlg.exec()
    
    def _enviar_alerta_email(self, items):
        
        # 🔹 CONFIGURACIÓN SMTP
        smtp_server = "smtp.gmail.com"
        smtp_port = 587
        smtp_user = "macacodrip121@gmail.com"      # Cambiá esto
        smtp_password = "12345"       # Cambiá esto

        # 🔹 DESTINATARIO
        destinatario = "macacodrip121@gmail.com"    # Cambiá esto si querés otro email

        asunto = "🔔 Alerta de Stock Bajo"
        cuerpo = "Los siguientes componentes están por debajo del stock mínimo:\n\n"
        cuerpo += "\n".join(
            f"- {r['nombre']} (Stock: {r['cantidad']}, Mínimo: {r['stock_minimo']})"
            for r in items
        )

        mensaje = MIMEText(cuerpo, "plain", "utf-8")
        mensaje["Subject"] = asunto
        mensaje["From"] = smtp_user
        mensaje["To"] = destinatario

        try:
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.sendmail(smtp_user, [destinatario], mensaje.as_string())
            server.quit()
            print("✅ Email de alerta enviado.")
        except Exception as e:
            print("❌ Error al enviar email:", e)
    
    def _aplicar_configuracion_visual(self):
        settings = QSettings("MiInventario", "AppStock")
        tema = settings.value("tema", "Claro")
        color = settings.value("color_acento", "#4F81BD")
        tam_fuente_str = settings.value("tam_fuente", "11pt")
        tam_fuente = int(tam_fuente_str.replace("pt", "").strip())

        # Cargar tema correspondiente
        if tema == "Claro":
            cargar_tema(self, "tema_claro.qss", color, f"{tam_fuente}pt")
        elif tema == "Oscuro":
            cargar_tema(self, "tema_oscuro.qss", color, f"{tam_fuente}pt")
        else:
            cargar_tema(self, "tema_intermedio.qss", color, f"{tam_fuente}pt")



    @Slot()
    def _abrir_configuracion(self):
        dlg = ConfiguracionDialog(self)
        if dlg.exec():
            self._aplicar_configuracion_visual()

    
    def _obtener_vista_stock(self):
        settings = QSettings("MiInventario", "AppStock")
        return settings.value("vista_stock", "detalle") 

    def _guardar_vista_stock(self, modo):
        settings = QSettings("MiInventario", "AppStock")
        settings.setValue("vista_stock", modo)






